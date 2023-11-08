import openpyxl
from datetime import datetime

def escolha_tag():
    print("\nVocê escolheu o cadastro de TAGS.\n\n")

    condominio_nome = input("\nInforme o nome do condominio.\n\n")
    numero_apto = input("\nInforme o número do apartamento.\n\n")
    quantidade_entrega = input("\nInforme a quantidade entregue.\n\n")
    
    modelo = input("\n Informe o modelo da tag entregue:\n\n1- MyFare\n2- R-fid\n3- Nice\n4- Intelbras\n\n")
    while modelo not in ["1","2","3","4"]:
            print("Opção inválida, reponda novamente.")
            modelo = input("\n Informe o modelo da tag entregue:\n\n1- MyFare\n2- R-fid\n3- Nice\n4- Intelbras\n\n")

    if modelo == "1":
            modelo = "MyFare"

    elif modelo ==  "2":
            modelo = "R-fid"  
        
    elif modelo ==  "3":
            modelo = "Nice" 

    elif modelo ==  "4":
           modelo  = "Intelbras" 

        
        
    metodo_pagamento = input("\nQual o método de pagamento?\n\n1- Pix\n2- Dinheiro\n3- DOC\n\n")

    while metodo_pagamento not in ["1","2","3"]:
        print("Opção inválida, reponda novamente.")
        metodo_pagamento = input("\nQual o método de pagamento?\n\n1- Pix\n2- Dinheiro\n3- DOC\n\n")

    if metodo_pagamento == "1":
            metodo_pagamento = "Pix"
        
    elif metodo_pagamento == "2":
            metodo_pagamento = "Dinheiro"

    elif metodo_pagamento == "3":
            metodo_pagamento = "DOC"
        
        
    metodo_entrega = input("\nComo foi entregue?\n\n1- Foi entregue\n2- Foi retirada\n\n")
    
    while metodo_entrega not in ["1","2","3"]:
            print("Opção inválida, reponda novamente.")
            metodo_entrega = input("\nComo foi entregue?\n\n1- Foi entregue\n2- Foi retirada\n\n")
        
    if metodo_entrega == "1":
            metodo_entrega = "Entregue"
        
    elif metodo_entrega =="2":
            metodo_entrega = "Retirada"
    
    data_entrega = input("Informe a data da entrega. Em formato DD-MM-AA\n\n")

    salvar_dados(condominio_nome, numero_apto, quantidade_entrega, modelo, metodo_pagamento, metodo_entrega, data_entrega)

    cadastrar_mais = input("Deseja cadastrar mais TAGS ou CONTROLES?\n\n1- Sim, quero cadastrar mais TAGS\n2- Sim, quero cadastrar mais CONTROLES\n3- Não, quero sair")

    while cadastrar_mais not in ["1","2","3"]:
        print("Opção inválida, escolha novamente.")
        cadastrar_mais

        if cadastrar_mais == "1":
            escolha_tag()
    
        elif cadastrar_mais == "2":
            escolha_controle()

        else:
             exit


def escolha_controle():    
    print("Você escolheu o cadastro de controles.\n\n")
    condominio_nome = input("\nInforme o nome do condominio.\n\n")
    numero_apto = input("\nInforme o número do apartamento.\n\n")
        #####
    quantidade_entrega = input("\nInforme a quantidade entregue.\n\n")
        ####
    modelo = input("\n Informe o modelo do controle entregue:\n\n1- CS\n2- Nice\n3- Adesivo placa Intelbras\n4- Adesivo parabrisa Intelbras\n5- Adesivo parabrisa Control ID\n\n")
    while modelo not in["1","2","3","4", "5"]:
        print("Opção inválida, responda novamente.")
        modelo
        
    if modelo == "1":
        modelo = "CS"
        
    elif modelo == "2":
        modelo = "Nice"
        
    elif modelo == "3":
        modelo = " Adesivo de placa Intelbras"
        
    elif modelo == "4":
        modelo = "Adesivo de parabrisa Intelbras"

    elif modelo == "5":
        modelo = "Adesivo de parabrisa Control ID"
        
        
    metodo_pagamento = input("\nQual o método de pagamento?\n\n1- Pix\n2- Dinheiro\n3- DOC\n\n")
    while metodo_pagamento not in ["1","2","3"]:
        print("Opção inválida, reponda novamente.")
        metodo_pagamento

    if metodo_pagamento == "1":
        metodo_pagamento = "Pix"
        
    elif metodo_pagamento == "2":
        metodo_pagamento = "Dinheiro"

    elif metodo_pagamento == "3":
        metodo_pagamento = "DOC"
 
        
    metodo_entrega = input("\nComo foi entregue?\n\n1- Foi entregue\n2- Foi retirada\n\n")
    while metodo_entrega not in ["1","2","3"]:
        print("Opção inválida, reponda novamente.")
        metodo_entrega
        
    if metodo_entrega == "1":
        metodo_entrega = "Entregue"
        
    elif metodo_entrega =="2":
        metodo_entrega = "Retirada"
        
    data_entrega = input("Informe a data da entrega. Em formato DD-MM-AA\n\n")

    salvar_dados(condominio_nome, numero_apto, quantidade_entrega, modelo, metodo_pagamento, metodo_entrega, data_entrega)

    cadastrar_mais = input("Deseja cadastrar mais TAGS ou CONTROLES?\n\n1- Sim, quero cadastrar mais TAGS\n2- Sim, quero cadastrar mais CONTROLES\n3- Não, quero sair\n\n")

    while cadastrar_mais not in ["1","2","3"]:
        print("Opção inválida, escolha novamente.")
        cadastrar_mais

    if cadastrar_mais == "1":
        escolha_tag()
    
    elif cadastrar_mais == "2":
        escolha_controle()

    else:
         exit


def salvar_dados(condominio_nome, numero_apto, quantidade_entrega, modelo, metodo_pagamento, metodo_entrega, data_entrega):
    try:
        workbook = openpyxl.load_workbook("registros_entregas.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet("ENTREGAS")
        workbook.save("registros_entregas.xlsx")

    sheet = workbook["ENTREGAS"]

    linha = 3
    while sheet.cell(row=linha, column=1).value is not None:
        linha += 1

    inputs = [condominio_nome, numero_apto, quantidade_entrega, modelo, metodo_pagamento, metodo_entrega, data_entrega]

    for coluna, valor in enumerate(inputs, start=1):
        sheet.cell(row=linha, column=coluna, value=valor)

    workbook.save("registros_entregas.xlsx")
    workbook.close()




print("Seja bem vindo.")
print("O que deseja fazer?\n")

escolha_menu = input("1- Registro de entrega de TAG\n2- Registro de entraga de CONTROLE\n\n")

if escolha_menu == "1":
    escolha_tag()


elif escolha_menu == "2":
    escolha_controle()